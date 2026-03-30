#!/usr/bin/env python3
"""
generate_newsletter.py

Reads an Excel file (columns: Type, Data, Title, Creator, Image) and generates
a newsletter EML file. HTML is produced with newsletter_renderer.render_newsletter()
(same stack as the Flask builder).

Section mapping (Type column values):
  Month News    → bullet list ("What's Going On")
  Save the Date → event list
  Product       → product cards (when block order includes General Information)
  General       → text blocks

Usage:
    python generate_newsletter.py
    python generate_newsletter.py --xlsx data.xlsx --out output.eml --month "April"
"""

import argparse
import base64
import copy
import html as html_module
import sys
import uuid
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl

from newsletter_renderer import render_newsletter

# ---------------------------------------------------------------------------
# Defaults and Configuration
# ---------------------------------------------------------------------------
DEFAULT_XLSX = "EmailData (2).xlsx"
DEFAULT_OUT  = "newsletter_output.eml"
DEFAULT_MONTH = "March"

# Email configuration with styling defaults
EMAIL_CONFIG = {
    "from": "ADIA EMEA Newsletter <newsletter@adia-emea.com>",
    "to": "adia@adia.com",
    "subject": "ADIA EMEA - Good to Know | {month} {year}",
    "colors": {
        "background": "#EEF6F7",
        "white": "#ffffff",
        "border": "#dddddd",
        "red_accent": "#cc0000",
        "text_dark": "#1a1a1a",
        "text_medium": "#444444",
        "text_light": "#555555",
        "text_muted": "#888888",
        "text_footer": "#aaaaaa",
        "divider": "#eeeeee",
        "banner_grey": "#888888",
        "header_bg": "#f9f9f9",
        "save_date_bg": "#E5EFF0",
        "footer_bg": "#1a1a1a",
    },
    "fonts": {
        "family": "Arial,Helvetica,sans-serif",
    },
    "sizes": {
        "table_width": "700",
        "banner_width": "494",
        "banner_height": "120",
    }
}

IMAGE_MIME = {
    ".png": "png", ".jpg": "jpeg", ".jpeg": "jpeg",
    ".gif": "gif", ".bmp": "bmp", ".webp": "webp",
}

# ---------------------------------------------------------------------------
# Read Excel
# ---------------------------------------------------------------------------

def read_excel_rows(xlsx_path: Path) -> dict[str, list[dict]]:
    """Return rows grouped by the Type column value."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Build a case-insensitive column-name → zero-based index map
    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = cell.column - 1

    def col(row_values, name: str):
        idx = headers.get(name.lower())
        return row_values[idx] if idx is not None and idx < len(row_values) else None

    grouped: dict[str, list[dict]] = {}
    for row_values in ws.iter_rows(min_row=2, values_only=True):
        type_val = col(row_values, "type")
        data     = col(row_values, "data")
        title    = col(row_values, "title")
        creator  = col(row_values, "creator")
        image    = col(row_values, "image")

        if not type_val and not data and not title:
            continue  # skip blank rows

        type_key = str(type_val).strip() if type_val else "General"
        grouped.setdefault(type_key, []).append({
            "data":    str(data).strip()    if data    else "",
            "title":   str(title).strip()   if title   else "",
            "creator": str(creator).strip() if creator else "",
            "image":   str(image).strip()   if image   else None,
        })

    return grouped


# ---------------------------------------------------------------------------
# Excel → newsletter_renderer config (same shape as Flask / static builder)
# ---------------------------------------------------------------------------

_DEFAULT_EXCEL_LAYOUT = [
    "Month News",
    "Save the Date",
    "General Information",
    "General",
]


def excel_to_newsletter_config(
    grouped: dict[str, list[dict]],
    *,
    email_config: dict | None = None,
    ordered_blocks: list[str] | None = None,
    block_bg_colors: dict[str, str] | None = None,
    meta: dict | None = None,
    bullet_heading: str | None = None,
) -> dict:
    """
    Build a newsletter config dict for newsletter_renderer.render_newsletter().

    ``ordered_blocks`` uses Streamlit/legacy block ids: Month News, Save the Date,
    General Information (product rows), General (text blocks). Header and footer
    are always included in the same style as the Flask builder.
    """
    cfg = email_config or EMAIL_CONFIG
    colors = cfg["colors"]
    fonts = cfg["fonts"]
    sizes = cfg["sizes"]
    meta = dict(meta) if meta else {}
    org = meta.get("orgName", "Aon")
    tagline = meta.get("tagline", "Newsletter")
    block_bg = dict(block_bg_colors) if block_bg_colors else {}

    tw = sizes.get("table_width", "700")
    try:
        table_width = int(str(tw).replace("px", ""))
    except ValueError:
        table_width = 700

    theme = {
        "primaryColor": colors.get("red_accent", "#E31837"),
        "backgroundColor": colors.get("background", "#F5F5F5"),
        "fontFamily": fonts.get("family", "Arial,Helvetica,sans-serif"),
        "tableWidth": table_width,
    }

    sections: list[dict] = []
    sections.append({
        "id": str(uuid.uuid4()),
        "type": "header",
        "props": {
            "orgName": org,
            "tagline": tagline,
            "backgroundColor": "#E31837",
            "textColor": "#ffffff",
        },
    })

    layout = list(ordered_blocks) if ordered_blocks else list(_DEFAULT_EXCEL_LAYOUT)
    bullet_title = bullet_heading if bullet_heading is not None else "What's Going On"
    bullet_color = theme["primaryColor"]

    for block_id in layout:
        if block_id == "Month News":
            month_rows = grouped.get("Month News", [])
            items = [r.get("data", "") for r in month_rows if str(r.get("data", "")).strip()]
            if not items:
                continue
            sections.append({
                "id": str(uuid.uuid4()),
                "type": "bullet_list",
                "props": {
                    "heading": bullet_title,
                    "items": items,
                    "bulletColor": bullet_color,
                    "backgroundColor": block_bg.get("Month News", colors.get("white", "#ffffff")),
                },
            })

        elif block_id == "Save the Date":
            std_rows = grouped.get("Save the Date", [])
            items = [r.get("data", "") for r in std_rows if str(r.get("data", "")).strip()]
            if not items:
                continue
            sections.append({
                "id": str(uuid.uuid4()),
                "type": "event_list",
                "props": {
                    "heading": "Save the Date!",
                    "items": items,
                    "backgroundColor": block_bg.get(
                        "Save the Date",
                        colors.get("save_date_bg", "#EEF6F7"),
                    ),
                },
            })

        elif block_id == "General Information":
            prod_bg = block_bg.get("General Information", colors.get("white", "#ffffff"))
            for row in grouped.get("Product", []):
                img = row.get("image") or ""
                sections.append({
                    "id": str(uuid.uuid4()),
                    "type": "product_card",
                    "props": {
                        "title": row.get("title", ""),
                        "body": row.get("data", ""),
                        "creator": row.get("creator", ""),
                        "imageUrl": str(img).strip() if img else "",
                        "backgroundColor": prod_bg,
                    },
                })

        elif block_id == "General":
            gen_bg = block_bg.get("General", colors.get("white", "#ffffff"))
            for row in grouped.get("General", []):
                sections.append({
                    "id": str(uuid.uuid4()),
                    "type": "text_block",
                    "props": {
                        "heading": row.get("title", ""),
                        "body": row.get("data", ""),
                        "backgroundColor": gen_bg,
                    },
                })

    sections.append({
        "id": str(uuid.uuid4()),
        "type": "footer",
        "props": {
            "orgName": org,
            "year": str(datetime.now().year),
            "backgroundColor": "#1A1A1A",
            "textColor": "#aaaaaa",
        },
    })

    return {"meta": meta, "theme": theme, "sections": sections}


def embed_local_images_in_config(config: dict, base_dir: Path) -> dict:
    """Return a deep copy with local imageUrl/logoUrl paths replaced by data URIs."""
    _mime = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".bmp": "image/bmp",
        ".webp": "image/webp",
    }
    out = copy.deepcopy(config)
    for section in out.get("sections", []):
        props = section.get("props", {})
        for key in ("imageUrl", "logoUrl"):
            url = props.get(key, "")
            if not url or str(url).startswith(("data:", "http://", "https://", "/")):
                continue
            path = Path(url)
            if not path.is_absolute():
                path = base_dir / url
            if path.is_file():
                mime = _mime.get(path.suffix.lower(), "image/png")
                b64 = base64.b64encode(path.read_bytes()).decode()
                props[key] = f"data:{mime};base64,{b64}"
    return out


# ---------------------------------------------------------------------------
# HTML row / block generators
# ---------------------------------------------------------------------------

def _month_news_row(data: str) -> str:
    e = html_module.escape(data)
    return (
        "              <tr>\n"
        "                <td valign=\"top\" width=\"14\" style=\"padding-top:3px;\">\n"
        "                  <div style=\"width:6px;height:6px;background-color:#cc0000;"
        "border-radius:50%;margin-top:5px;\"></div>\n"
        "                </td>\n"
        "                <td style=\"padding:0 0 14px 8px;\">\n"
        f"                  <div style=\"font-size:13px;color:#1a1a1a;font-weight:bold;"
        f"line-height:1.5;\">{e}</div>\n"
        "                </td>\n"
        "              </tr>"
    )


def _save_the_date_row(data: str) -> str:
    e = html_module.escape(data)
    return (
        "              <tr>\n"
        "                <td valign=\"top\" width=\"14\">\n"
        "                  <div style=\"width:6px;height:6px;background-color:#1a1a1a;"
        "border-radius:50%;margin-top:5px;\"></div>\n"
        "                </td>\n"
        "                <td style=\"padding:0 0 8px 8px;\">\n"
        f"                  <div style=\"font-size:13px;color:#1a1a1a;"
        f"line-height:1.5;\">{e}</div>\n"
        "                </td>\n"
        "              </tr>"
    )


def _product_block(title: str, data: str, creator: str,
                   image_cid: str | None, background_color: str | None = None) -> str:
    t = html_module.escape(title)
    d = html_module.escape(data)
    bg_style = f"background-color:{background_color};" if background_color else ""
    img_html = (
        f"            <img src=\"cid:{image_cid}\" alt=\"{t}\" "
        "style=\"display:block;max-width:100%;height:auto;margin-bottom:10px;\" />\n"
        if image_cid else ""
    )
    creator_html = (
        f"            <div style=\"font-size:11px;color:#888888;margin-top:8px;"
        f"font-style:italic;\">By {html_module.escape(creator)}</div>\n"
        if creator else ""
    )
    return (
        "        <tr>\n"
        f"          <td style=\"padding:16px 28px 8px 28px;{bg_style}\">\n"
        f"            <div style=\"font-size:16px;font-weight:bold;color:#1a1a1a;"
        f"margin-bottom:8px;\">{t}</div>\n"
        + img_html
        + f"            <div style=\"font-size:13px;color:#444444;"
          f"line-height:1.6;\">{d}</div>\n"
        + creator_html
        + "          </td>\n"
          "        </tr>\n"
          f"        <tr><td style=\"height:1px;background-color:#eeeeee;{bg_style}"
          "font-size:0;line-height:0;\">&nbsp;</td></tr>"
    )


def _general_block(title: str, data: str, is_last: bool, background_color: str | None = None) -> str:
    t = html_module.escape(title)
    d = html_module.escape(data)
    pad = "22px 28px 28px 28px" if is_last else "22px 28px 6px 28px"
    bg_style = f"background-color:{background_color};" if background_color else ""
    return (
        "        <tr>\n"
        f"          <td style=\"padding:{pad};{bg_style}\">\n"
        f"            <div style=\"font-size:14px;font-weight:bold;color:#1a1a1a;"
        f"margin-bottom:8px;\">{t}</div>\n"
        f"            <div style=\"font-size:13px;color:#444444;"
        f"line-height:1.6;\">{d}</div>\n"
        "          </td>\n"
        "        </tr>"
    )


_GENERAL_DIVIDER = (
    "        <tr><td style=\"height:18px;font-size:0;line-height:0;\">&nbsp;</td></tr>\n"
    "        <tr><td style=\"height:1px;background-color:#eeeeee;"
    "font-size:0;line-height:0;\">&nbsp;</td></tr>"
)


# ---------------------------------------------------------------------------
# Dynamic HTML Builder
# ---------------------------------------------------------------------------

DEFAULT_LAYOUT: list[str] = [
    "Header",
    "Month News",
    "Save the Date",
    "General Information",
    "General",
    "Footer",
]


def _resolve_block_bg(block_id: str, config: dict, block_bg_colors: dict[str, str] | None) -> str | None:
    """
    Return the background color for a block type.
    If block_bg_colors provides an override, it wins; otherwise defaults are used.
    """
    colors = config["colors"]
    defaults: dict[str, str] = {
        "Month News": colors.get("white", "#ffffff"),
        "Save the Date": colors.get("save_date_bg", "#E5EFF0"),
        "General Information": colors.get("white", "#ffffff"),
        "General": colors.get("white", "#ffffff"),
    }
    if block_bg_colors and block_id in block_bg_colors and block_bg_colors[block_id]:
        return block_bg_colors[block_id]
    return defaults.get(block_id)


def _build_header_section(config: dict) -> str:
    """Build the header section with ADIA EMEA branding and grey banner."""
    colors = config["colors"]
    sizes = config["sizes"]
    return (
        "        <!-- HEADER BANNER -->\n"
        "        <tr>\n"
        "          <td style=\"padding:0;background-color:#ffffff;\">\n"
        "            <table width=\"100%\" cellpadding=\"0\" cellspacing=\"0\" border=\"0\">\n"
        "              <tr>\n"
        "                <!-- Red accent bar -->\n"
        f"                <td width=\"6\" style=\"background-color:{colors['red_accent']};\">&nbsp;</td>\n"
        "                <!-- Title area -->\n"
        "                <td width=\"200\" valign=\"middle\" style=\"padding:18px 16px 18px 14px;"
        f"background-color:{colors['header_bg']};\">\n"
        "                  <div style=\"font-size:22px;font-weight:bold;color:#1a1a1a;"
        "letter-spacing:0.5px;line-height:1.2;\">ADIA EMEA</div>\n"
        "                  <div style=\"font-size:13px;color:#555555;margin-top:4px;"
        "font-style:italic;\">Good to Know</div>\n"
        "                </td>\n"
        "                <!-- Header banner (grey block) -->\n"
        f"                <td style=\"padding:0;vertical-align:top;background-color:{colors['banner_grey']};"
        f"width:{sizes['banner_width']}px;height:{sizes['banner_height']}px;\"></td>\n"
        "              </tr>\n"
        "            </table>\n"
        "          </td>\n"
        "        </tr>\n"
        "\n"
        "        <!-- DIVIDER -->\n"
        f"        <tr><td style=\"height:1px;background-color:{colors['border']};"
        "font-size:0;line-height:0;\">&nbsp;</td></tr>"
    )


def _build_footer_section(config: dict) -> str:
    """Build the footer section with copyright and links."""
    colors = config["colors"]
    current_year = datetime.now().year
    return (
        "\n"
        "        <!-- FOOTER -->\n"
        "        <tr>\n"
        f"          <td style=\"padding:16px 28px;background-color:{colors['footer_bg']};text-align:center;\">\n"
        f"            <div style=\"font-size:11px;color:{colors['text_footer']};line-height:1.5;\">"
        f"&copy; {current_year} ADIA EMEA &bull; All rights reserved</div>\n"
        f"            <div style=\"font-size:11px;color:{colors['text_muted']};margin-top:4px;\">\n"
        f"              <a href=\"#\" style=\"color:{colors['red_accent']};text-decoration:none;\">"
        "Unsubscribe</a>\n"
        "              &nbsp;&bull;&nbsp;\n"
        f"              <a href=\"#\" style=\"color:{colors['text_footer']};text-decoration:none;\">"
        "View in browser</a>\n"
        "            </div>\n"
        "          </td>\n"
        "        </tr>"
    )


def build_html_email(grouped_data: dict[str, list[dict]], month_name: str,
                     config: dict, image_cids: dict[str, str],
                     layout: list[str] | None = None,
                     block_bg_colors: dict[str, str] | None = None) -> str:
    """
    Build the complete HTML email structure dynamically.
    
    Args:
        grouped_data: Dictionary mapping Type values to lists of row data
        month_name: Name of the month for the "What's going on in [Month]" section
        config: Email configuration dictionary with styling options
        image_cids: Dictionary mapping image paths to Content-IDs
    
    Returns:
        Complete HTML string for the email
    """
    colors = config["colors"]
    fonts = config["fonts"]
    sizes = config["sizes"]

    layout_to_render = layout[:] if layout else DEFAULT_LAYOUT[:]

    def render_block(block_id: str) -> str:
        if block_id == "Header":
            return _build_header_section(config)
        if block_id == "Footer":
            return _build_footer_section(config)

        bg = _resolve_block_bg(block_id, config, block_bg_colors)
        bg_style = f"background-color:{bg};" if bg else ""

        if block_id == "Month News":
            month_rows = grouped_data.get("Month News", [])
            if not month_rows:
                return ""
            rows_html = "\n".join(_month_news_row(r["data"]) for r in month_rows)
            return (
                "\n"
                f"        <!-- WHAT'S GOING ON IN {month_name.upper()} -->\n"
                "        <tr>\n"
                f"          <td style=\"padding:24px 28px 8px 28px;{bg_style}\">\n"
                f"            <div style=\"font-size:16px;font-weight:bold;color:#1a1a1a;"
                f"margin-bottom:14px;padding-bottom:6px;border-bottom:1px solid {colors['divider']};\">"
                f"What&#39;s going on in {html_module.escape(month_name)}...</div>\n"
                "            <table width=\"100%\" cellpadding=\"0\" cellspacing=\"0\" border=\"0\">\n"
                + rows_html + "\n"
                "            </table>\n"
                "          </td>\n"
                "        </tr>\n"
                "\n"
                f"        <!-- DIVIDER -->\n"
                f"        <tr><td style=\"height:1px;background-color:{colors['divider']};{bg_style}"
                "font-size:0;line-height:0;\">&nbsp;</td></tr>"
            )

        if block_id == "Save the Date":
            std_rows = grouped_data.get("Save the Date", [])
            if not std_rows:
                return ""
            rows_html = "\n".join(_save_the_date_row(r["data"]) for r in std_rows)
            return (
                "\n"
                "        <!-- SAVE THE DATE -->\n"
                "        <tr>\n"
                f"          <td style=\"padding:20px 28px 8px 28px;{bg_style}\">\n"
                f"            <div style=\"font-size:15px;font-weight:bold;color:{colors['red_accent']};"
                "margin-bottom:12px;\">Save the Date!</div>\n"
                "            <table width=\"100%\" cellpadding=\"0\" cellspacing=\"0\" border=\"0\">\n"
                + rows_html + "\n"
                "            </table>\n"
                "          </td>\n"
                "        </tr>\n"
                "\n"
                f"        <!-- DIVIDER -->\n"
                f"        <tr><td style=\"height:1px;background-color:{colors['border']};{bg_style}"
                "font-size:0;line-height:0;\">&nbsp;</td></tr>"
            )

        if block_id == "General Information":
            product_rows = grouped_data.get("Product", [])
            if not product_rows:
                return ""
            blocks: list[str] = []
            for row in product_rows:
                image_path = row.get("image")
                cid = image_cids.get(image_path) if image_path else None
                blocks.append(_product_block(row["title"], row["data"], row["creator"], cid, bg))
            return (
                "\n"
                "        <!-- GENERAL INFORMATION HEADING -->\n"
                "        <tr>\n"
                f"          <td style=\"padding:28px 28px 10px 28px;{bg_style}\">\n"
                "            <div style=\"font-size:24px;font-weight:bold;color:#1a1a1a;"
                "letter-spacing:0.3px;\">General Information</div>\n"
                "          </td>\n"
                "        </tr>\n"
                "\n"
                "        <!-- GENERAL INFO DIVIDER LINE -->\n"
                f"        <tr><td style=\"height:2px;background-color:{colors['divider']};{bg_style}"
                "margin:0 28px;font-size:0;line-height:0;\">&nbsp;</td></tr>\n"
                "\n"
                "        <!-- GENERAL INFORMATION BODY -->\n"
                + "\n".join(blocks)
            )

        if block_id == "General":
            general_rows = grouped_data.get("General", [])
            if not general_rows:
                return ""
            blocks: list[str] = []
            for i, row in enumerate(general_rows):
                is_last = (i == len(general_rows) - 1)
                blocks.append(_general_block(row["title"], row["data"], is_last, bg))
                if not is_last:
                    blocks.append(_GENERAL_DIVIDER)
            return "\n" + "\n".join(blocks)

        return ""

    rendered_blocks = [render_block(block_id) for block_id in layout_to_render]
    rendered = "".join(b for b in rendered_blocks if b)
    
    # Assemble complete HTML
    html = (
        "<!DOCTYPE html>\n"
        "<html lang=\"en\">\n"
        "<head>\n"
        "<meta charset=\"UTF-8\">\n"
        "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">\n"
        "<title>ADIA EMEA - Good to Know</title>\n"
        "</head>\n"
        f"<body style=\"margin:0;padding:0;background-color:{colors['background']};"
        f"font-family:{fonts['family']};\">\n"
        f"<table width=\"100%\" cellpadding=\"0\" cellspacing=\"0\" border=\"0\" "
        f"style=\"background-color:{colors['background']};\">\n"
        "  <tr>\n"
        f"    <td align=\"center\" valign=\"top\" style=\"padding:20px 10px;background-color:{colors['background']};\">\n"
        f"      <table width=\"{sizes['table_width']}\" cellpadding=\"0\" cellspacing=\"0\" "
        f"border=\"0\" bgcolor=\"{colors['white']}\" style=\"background-color:{colors['white']} !important;border:1px solid {colors['border']};\">\n"
        + rendered
        + "\n"
        "      </table>\n"
        "    </td>\n"
        "  </tr>\n"
        "</table>\n"
        "</body>\n"
        "</html>"
    )
    return html


# ---------------------------------------------------------------------------
# Dynamic EML Builder
# ---------------------------------------------------------------------------

def build_eml_message(html_content: str, from_addr: str, to_addr: str,
                     subject: str, date_str: str = None) -> MIMEMultipart:
    """
    Build a multipart/related EML message structure.
    
    Args:
        html_content: HTML content for the email body
        from_addr: From email address
        to_addr: To email address
        subject: Email subject line
        date_str: Date string (defaults to current date)
    
    Returns:
        MIMEMultipart message ready for attachments
    """
    # Create multipart/related message
    msg = MIMEMultipart("related")
    
    # Set headers
    msg["MIME-Version"] = "1.0"
    if date_str:
        msg["Date"] = date_str
    else:
        msg["Date"] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0000")
    msg["Message-ID"] = f"<adia-emea-goodtoknow-{datetime.now().strftime('%Y%m%d')}@adia.emea>"
    msg["From"] = from_addr
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg["Content-Type"] = 'multipart/related; type="text/html"'
    
    # Create HTML part
    html_part = MIMEText(html_content, "html", "utf-8")
    html_part["Content-Transfer-Encoding"] = "base64"
    
    # Encode HTML as base64
    encoded = base64.encodebytes(html_content.encode("utf-8")).decode("ascii")
    html_part.set_payload(encoded)
    
    # Attach HTML part
    msg.attach(html_part)
    
    return msg


# ---------------------------------------------------------------------------
# Image attachment helper
# ---------------------------------------------------------------------------

def _load_image_part(img_value: str, base_dir: Path):
    """Load an image file and create a MIMEImage part with Content-ID."""
    from email.mime.image import MIMEImage
    
    img_path = Path(img_value)
    if not img_path.is_absolute():
        img_path = base_dir / img_value
    if not img_path.exists():
        print(f"Warning: image not found: {img_path}", file=sys.stderr)
        return None, None
    subtype = IMAGE_MIME.get(img_path.suffix.lower(), "png")
    cid = str(uuid.uuid4())
    with open(img_path, "rb") as fh:
        img_data = fh.read()
    part = MIMEImage(img_data, _subtype=subtype)
    part["Content-ID"] = f"<{cid}>"
    part["Content-Disposition"] = f'inline; filename="{img_path.name}"'
    part["Content-Description"] = img_path.name
    return part, cid


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate a newsletter EML file from an Excel data file."
    )
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to Excel data file")
    parser.add_argument("--out",  default=DEFAULT_OUT,  help="Output EML path")
    parser.add_argument("--month", default=DEFAULT_MONTH, help="Month name for the newsletter")
    parser.add_argument("--from", default=EMAIL_CONFIG["from"], help="From email address")
    parser.add_argument("--to", default=EMAIL_CONFIG["to"], help="To email address")
    parser.add_argument("--subject", default=None, help="Email subject (defaults to config template)")
    args = parser.parse_args()

    script_dir = Path(__file__).parent
    xlsx_path = Path(args.xlsx) if Path(args.xlsx).is_absolute() else script_dir / args.xlsx
    out_path  = Path(args.out)  if Path(args.out).is_absolute()  else script_dir / args.out

    # ------------------------------------------------------------------
    # 1. Read Excel data
    # ------------------------------------------------------------------
    grouped = read_excel_rows(xlsx_path)
    print(f"Loaded data from {xlsx_path.name}:")
    for type_key, rows in grouped.items():
        print(f"  {type_key}: {len(rows)} row(s)")

    # ------------------------------------------------------------------
    # 2. Prepare image attachments and collect CIDs
    # ------------------------------------------------------------------
    image_cids: dict[str, str] = {}  # Maps image path to CID
    image_parts: dict[str, object] = {}  # Maps image path to MIMEImage part

    # Collect all unique image paths from Product rows
    product_rows = grouped.get("Product", [])
    for row in product_rows:
        if row.get("image") and row["image"] not in image_cids:
            img_part, cid = _load_image_part(row["image"], xlsx_path.parent)
            if img_part is not None:
                image_cids[row["image"]] = cid
                image_parts[row["image"]] = img_part
                print(f"    Prepared image: {row['image']} (CID: {cid})")

    # ------------------------------------------------------------------
    # 3. Subject + config, then render HTML (newsletter_renderer)
    # ------------------------------------------------------------------
    if args.subject:
        subject = args.subject.strip()
    else:
        current_year = datetime.now().year
        subject = EMAIL_CONFIG["subject"].format(month=args.month, year=current_year)

    from_addr = getattr(args, "from")
    meta = {
        "newsletterName": "Newsletter",
        "subject": subject,
        "from": from_addr,
        "to": args.to,
        "orgName": "Aon",
        "tagline": "Newsletter",
    }
    config = excel_to_newsletter_config(
        grouped,
        email_config=EMAIL_CONFIG,
        ordered_blocks=None,
        block_bg_colors=None,
        meta=meta,
        bullet_heading="What's Going On",
    )
    html = render_newsletter(config, image_cids=image_cids)
    print(f"  Rendered HTML via newsletter_renderer ({len(grouped)} Excel section type(s)).")

    # ------------------------------------------------------------------
    # 4. Build EML message
    # ------------------------------------------------------------------
    msg = build_eml_message(html, from_addr, args.to, subject)

    # Attach images
    for image_path, img_part in image_parts.items():
        msg.attach(img_part)
        print(f"  Attached image: {image_path}")

    # ------------------------------------------------------------------
    # 5. Write output EML
    # ------------------------------------------------------------------
    with open(out_path, "wb") as fh:
        fh.write(msg.as_bytes())

    print(f"\nDone. Output written to: {out_path}")


if __name__ == "__main__":
    main()
