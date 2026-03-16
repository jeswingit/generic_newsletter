#!/usr/bin/env python3
"""
template_generator.py

Helper functions for generating Excel templates and other utilities.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


def create_excel_template() -> BytesIO:
    """
    Create a sample Excel template with the correct column structure and example data.
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Newsletter Data"
    
    # Define headers
    headers = ["Type", "Data", "Title", "Creator", "Image"]
    
    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Example data rows
    example_data = [
        # Month News examples
        ["Month News", "New product launch scheduled for next quarter", "", "", ""],
        ["Month News", "Team meeting scheduled for the 15th", "", "", ""],
        ["Month News", "Office hours updated - check the intranet for details", "", "", ""],
        # Save the Date examples
        ["Save the Date", "Annual Conference - March 25-27, 2026 - London", "", "", ""],
        ["Save the Date", "Product Demo Webinar - April 10, 2026 at 2:00 PM", "", "", ""],
        # Product examples
        ["Product", "Our latest product offers enhanced features and improved performance. It's designed to streamline your workflow and increase productivity.", "New Product Release", "John Smith", "images/product1.jpg"],
        ["Product", "This innovative solution addresses key challenges in the industry. Learn more about how it can benefit your organization.", "Innovation Spotlight", "Jane Doe", "images/product2.jpg"],
        # General examples
        ["General", "This section contains general information and updates that don't fit into other categories.", "Important Update", "", ""],
        ["General", "Please review the new policies and procedures that have been updated this month.", "Policy Updates", "", ""],
    ]
    
    # Write example data
    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add instructions in a separate sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["Newsletter Excel Template - Instructions"],
        [""],
        ["Column Descriptions:"],
        ["Type", "Content type: 'Month News', 'Save the Date', 'Product', or 'General'"],
        ["Data", "Main content text for the newsletter item"],
        ["Title", "Title/heading (required for Product and General types)"],
        ["Creator", "Author name (optional, typically used for Product type)"],
        ["Image", "Path to image file (optional, typically used for Product type)"],
        [""],
        ["Content Type Guidelines:"],
        ["Month News", "Bullet list items for monthly updates. Only 'Data' column is required."],
        ["Save the Date", "Event announcements. Only 'Data' column is required."],
        ["Product", "Product spotlight cards. Requires 'Title', 'Data', and optionally 'Creator' and 'Image'."],
        ["General", "Informational blocks. Requires 'Title' and 'Data'."],
        [""],
        ["Tips:"],
        ["- Image paths should be relative to the Excel file location"],
        ["- Leave cells empty if not applicable"],
        ["- You can delete the example rows and add your own data"],
        ["- Column names are case-insensitive"],
    ]
    
    for row_idx, row_data in enumerate(instructions, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif isinstance(value, str) and value.endswith(":"):
                cell.font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 70
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
